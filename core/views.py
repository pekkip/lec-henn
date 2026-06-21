import io
import json
import logging
import random
import string
import re
from datetime import date, timedelta
from datetime import datetime as dt
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden

logger = logging.getLogger(__name__)

from .models import (
    Client, ContactClient, Devis, LigneDevis,
    Facture, LigneFacture, AuditLog, ProfilUtilisateur,
    Territoire, Service, Equipe, ParametresAssociation, Bibliotheque,
    BibliothequeAides, PALETTE_COULEURS,
)
from .permissions import (
    peut_modifier_devis, peut_supprimer_devis, peut_voir_devis,
    peut_valider_facture, peut_supprimer_facture,
    peut_voir_facture, peut_modifier_facture,
    is_admin,
    peut_gerer_utilisateurs, peut_gerer_cet_utilisateur,
    get_collegues_ids, peut_acceder_compta,
)
from .dashboard_widgets import resolve_dashboard, sanitize_config
from .totaux import attacher_totaux_devis
# ══════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════

def get_profil(user):
    """Retourne le profil de l'utilisateur, le crée si nécessaire."""
    profil, _ = ProfilUtilisateur.objects.get_or_create(user=user)
    return profil


def parse_json_request(request):
    """Décode request.body en JSON. Retourne (data, None) ou (None, JsonResponse 400)."""
    try:
        return json.loads(request.body), None
    except (json.JSONDecodeError, ValueError, TypeError):
        return None, JsonResponse({'ok': False, 'error': 'JSON invalide'}, status=400)


def json_error(message, status=400):
    return JsonResponse({'ok': False, 'error': message}, status=status)


def json_error_permission():
    return JsonResponse({'ok': False, 'error': 'Permission refusée'}, status=403)


def to_decimal(val, default=None):
    if val is None or val == '':
        return default
    try:
        return Decimal(str(val))
    except (InvalidOperation, ValueError, TypeError):
        return default


def paginer(request, queryset, par_page=50):
    """Pagine un queryset et renvoie (page_obj, base_qs).

    `base_qs` = la query string courante SANS le paramètre `page`, pour
    construire des liens de pagination qui conservent les filtres actifs.
    """
    page_obj = Paginator(queryset, par_page).get_page(request.GET.get('page'))
    params = request.GET.copy()
    params.pop('page', None)
    return page_obj, params.urlencode()


def gen_numero_sequence(prefix, model, field, queryset=None):
    """Génère le prochain numéro libre de la forme {prefix}-{annee}-NNN pour
    model.field. Par défaut, scanne les valeurs commençant par '{prefix}-{annee}-'
    et prend max+1. Un `queryset` explicite permet de DÉCOUPLER la séquence de scan
    du préfixe affiché (cf. NUMEROTATION_FACTURE : compteur partagé entre type_docs).

    Race condition max+1 connue (dette technique, hors scope — proba nulle à
    l'échelle beta)."""
    year = timezone.localdate().year
    if queryset is None:
        queryset = model.objects.filter(
            **{f'{field}__startswith': f'{prefix}-{year}-'})
    nums = []
    for val in queryset.values_list(field, flat=True):
        tail = (val or '').split('-')[-1]
        if tail.isdigit():
            nums.append(int(tail))
    next_num = max(nums) + 1 if nums else 1
    return f"{prefix}-{year}-{str(next_num).zfill(3)}"


# Numérotation des factures : découple le préfixe AFFICHÉ de la SÉQUENCE de comptage.
# Pour basculer les appels sur la séquence FAC (en gardant le préfixe APP), il suffit
# de passer 'appel' -> {'prefix': 'APP', 'sequence': 'FAC'} (une ligne).
# ⚠️ Stratégie de numérotation (séquences partagées vs séparées, format) = point
# d'incertitude : arbitrage légal à venir (direction + conseil de l'association).
# Voir NOTES_DEV § Dette technique. Tant que ce n'est pas tranché, on conserve le
# découplage prefix/sequence (gen_numero_facture scanne par groupe de séquence).
NUMEROTATION_FACTURE = {
    'facture':   {'prefix': 'FAC', 'sequence': 'FAC'},
    'acompte':   {'prefix': 'FAC', 'sequence': 'FAC'},
    'structure': {'prefix': 'FAC', 'sequence': 'FAC'},  # partage la séquence FAC
    'appel':     {'prefix': 'APP', 'sequence': 'APP'},  # séquence propre (pour l'instant)
    'avoir':     {'prefix': 'AV',  'sequence': 'AV'},
}


def gen_numero_facture(type_doc):
    """Génère le prochain numéro de facture selon le type_doc (préfixe + séquence).
    Le scan porte sur tous les type_docs partageant la séquence (compteur commun),
    l'affichage utilise le préfixe — préserve le découplage NUMEROTATION_FACTURE."""
    cfg = NUMEROTATION_FACTURE.get(type_doc, NUMEROTATION_FACTURE['facture'])
    group = [td for td, c in NUMEROTATION_FACTURE.items() if c['sequence'] == cfg['sequence']]
    qs = Facture.objects.filter(type_doc__in=group, numero__isnull=False)
    return gen_numero_sequence(cfg['prefix'], Facture, 'numero', queryset=qs)


def add_audit(user, action, devis=None, facture=None, bypass=False):
    AuditLog.objects.create(
        user=user, action=action,
        devis=devis, facture=facture, bypass=bypass
    )


def ligne_to_dict(ligne):
    mo = ligne.total_mo()
    mat = ligne.total_mat()
    return {
        'id': ligne.pk,
        'type_ligne': ligne.type_ligne,
        'description': ligne.description,
        'quantite': float(ligne.quantite),
        'unite': ligne.unite,
        'cout_unitaire': float(ligne.cout_unitaire) if ligne.cout_unitaire is not None else None,
        'cout_mo': float(mo) if mo is not None else None,
        'cout_mat': float(mat) if mat is not None else None,
        'total': float(ligne.total()),
        'ordre': ligne.ordre,
        'ouvert': ligne.ouvert,
        'parent_id': ligne.parent_id,
        'aide_id': ligne.aide_id,
        'enfants': [ligne_to_dict(e) for e in ligne.enfants.all()],
    }


def ligne_facture_to_dict(ligne, deja_par_source=None):
    """
    Sérialise une LigneFacture en dict JSON pour le frontend.

    deja_par_source : dict {ligne_devis_source_id: {'montant', 'qty', 'refs'}}
    Calculé une fois pour toute la facture et passé en paramètre (évite N+1).
    """
    if deja_par_source is None:
        deja_par_source = {}

    source_id = ligne.ligne_devis_source_id
    entry = deja_par_source.get(source_id) if source_id else None
    deja_montant = entry['montant'] if entry else 0.0
    refs_deja = [{'ref': r, 'montant': m} for r, m in entry['refs'].items()] if entry else []

    return {
        'id': ligne.pk,
        'type_ligne': ligne.type_ligne,
        'description': ligne.description,
        'quantite': float(ligne.quantite),
        'quantite_originale': float(ligne.quantite_originale),
        'unite': ligne.unite,
        'cout_unitaire': float(ligne.cout_unitaire) if ligne.cout_unitaire is not None else None,
        'ordre': ligne.ordre,
        'ouvert': ligne.ouvert,
        'parent_id': ligne.parent_id,
        'deja_facture': deja_montant,
        'refs_deja_facture': refs_deja,
        'ligne_devis_source_id': source_id,
        'enfants': [ligne_facture_to_dict(e, deja_par_source) for e in ligne.enfants.all()],
    }


def build_lignes_creator(model, fk_kwargs, *, with_ouvert=True, with_aide=False,
                         with_quantite_originale=False, with_source=False,
                         recurse_only_titre=False):
    """Construit le récréateur d'arbre de lignes depuis le JSON du frontend,
    partagé par les 3 éditeurs (devis / facture / facture compta).

    fk_kwargs : {'devis': devis} ou {'facture': facture}. Les flags activent les
    champs spécifiques à chaque éditeur (cf. Phase 5, docs/plan_ameliorations.md) :
    - with_ouvert            : champ `ouvert` (devis + facture ; pas la compta).
    - with_aide              : résout `aide_id` → BibliothequeAides (devis seul).
    - with_quantite_originale: snapshot qty devis avec fallback sur `quantite` (facture).
    - with_source            : `ligne_devis_source_id` (facture, lien pré-remplissage).
    - recurse_only_titre     : ne descend dans les enfants que sous un TITRE (compta).
    """
    def create_lignes(items, parent=None, ordre=0):
        for item in items:
            type_ligne = item.get('type_ligne', 'F')
            fields = dict(
                **fk_kwargs, parent=parent, type_ligne=type_ligne,
                description=item.get('description', ''),
                quantite=to_decimal(item.get('quantite'), default=Decimal('1')),
                unite=item.get('unite', ''),
                cout_unitaire=to_decimal(item.get('cout_unitaire')),
                ordre=ordre,
            )
            if with_ouvert:
                fields['ouvert'] = item.get('ouvert', True)
            if with_aide:
                aide_id = item.get('aide_id')
                fields['aide'] = (
                    BibliothequeAides.objects.filter(pk=aide_id).first()
                    if aide_id else None
                )
            if with_quantite_originale:
                fields['quantite_originale'] = to_decimal(
                    item.get('quantite_originale', item.get('quantite')),
                    default=Decimal('1'))
            if with_source:
                fields['ligne_devis_source_id'] = item.get('ligne_devis_source_id')
            ligne = model.objects.create(**fields)
            if not recurse_only_titre or type_ligne == 'TITRE':
                create_lignes(item.get('enfants', []), parent=ligne)
            ordre += 1
    return create_lignes


# Types dont la quantité reflète le prix unitaire et non le métrage facturé.
# On ne leur applique jamais de réduction deja_par_source.
_TYPES_STRUCTURELS = frozenset({'OUV', 'MO', 'MAT'})


def copier_lignes_devis_vers_facture(lignes_devis, facture, parent_facture=None, ordre=0,
                                      deja_par_source=None, keep_qty=False):
    """
    Copie récursivement les lignes du devis vers la facture.

    Règles de pré-remplissage (si deja_par_source fourni) :
    - TITRE : démarre à qty devis ; se replie à 0 si et seulement si
      il était inclus dans une facture précédente ET que son total() == 0
      (tous les enfants « métrage » sont épuisés).
    - Seuls les **postes facturables de premier niveau** (enfants directs d'un TITRE,
      ou racines sans TITRE — C, S, F, FMO, FMAT, FIN, FINX…) sont réduits :
      max(0, qty_devis - qty_deja). Tout ce qui est **sous** un tel poste (la recette
      unitaire de la composite/section : OUV, MO, MAT, mais aussi une éventuelle
      sous-section S/C) conserve sa qty devis → drapeau `keep_qty=True` propagé en
      descendant. Sans ça, une section facturée une fois tombait à 0 et annulait le
      total de sa composite parente (et donc du TITRE).
    - _TYPES_STRUCTURELS (OUV, MO, MAT) : jamais réduits non plus (prix unitaire),
      garde-fou conservé même au premier niveau.
    quantite_originale garde la qty devis figée comme référence (snapshot).
    """
    for ligne in lignes_devis:
        if ligne.type_ligne == 'TITRE':
            entry = deja_par_source.get(ligne.pk) if deja_par_source else None
            was_billed = bool(entry and entry['qty'] > 0)

            lf = LigneFacture.objects.create(
                facture=facture,
                parent=parent_facture,
                type_ligne='TITRE',
                description=ligne.description,
                quantite=ligne.quantite,
                quantite_originale=ligne.quantite,
                unite=ligne.unite,
                cout_unitaire=ligne.cout_unitaire,
                ordre=ordre,
                ouvert=ligne.ouvert,
                ligne_devis_source=ligne,
            )
            # Les enfants d'un TITRE sont des postes facturables → keep_qty inchangé.
            copier_lignes_devis_vers_facture(
                ligne.enfants.all(), facture, lf, 0, deja_par_source, keep_qty=keep_qty
            )
            if was_billed and lf.total() == Decimal('0'):
                lf.quantite = Decimal('0')
                lf.save(update_fields=['quantite'])
        else:
            reductible = (deja_par_source is not None and not keep_qty
                          and ligne.type_ligne not in _TYPES_STRUCTURELS)
            if reductible:
                entry = deja_par_source.get(ligne.pk)
                qty_deja = entry['qty'] if entry else Decimal('0')
            else:
                qty_deja = Decimal('0')
            qty_nouvelle = max(Decimal('0'), ligne.quantite - qty_deja)

            lf = LigneFacture.objects.create(
                facture=facture,
                parent=parent_facture,
                type_ligne=ligne.type_ligne,
                description=ligne.description,
                quantite=qty_nouvelle,
                quantite_originale=ligne.quantite,
                unite=ligne.unite,
                cout_unitaire=ligne.cout_unitaire,
                ordre=ordre,
                ouvert=ligne.ouvert,
                ligne_devis_source=ligne,
            )
            if ligne.enfants.exists():
                # On est entré dans un poste facturable : tout en dessous est la
                # recette unitaire → conserver la qty devis (keep_qty=True).
                copier_lignes_devis_vers_facture(
                    ligne.enfants.all(), facture, lf, 0, deja_par_source, keep_qty=True
                )
        ordre += 1


# ══════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════

def login_view(request):
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    error = None
    if request.method == 'POST':
        user = authenticate(
            request,
            username=request.POST.get('username'),
            password=request.POST.get('password')
        )
        if user:
            login(request, user)
            return redirect('core:dashboard')
        error = "Identifiants incorrects."
    return render(request, 'core/login.html', {'error': error})


def logout_view(request):
    logout(request)
    return redirect('core:login')

def aide_view(request):
    return render(request, 'core/aide.html', {'site_url': settings.SITE_URL})

DOMAINE_AUTORISE = 'compagnonsbatisseurs.eu'

def mot_de_passe_oublie(request):
    succes = False
    erreur = None

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()

        if not email.endswith(f'@{DOMAINE_AUTORISE}'):
            erreur = f'Seules les adresses @{DOMAINE_AUTORISE} sont acceptées.'
        else:
            user = User.objects.filter(email__iexact=email, is_active=True).first()
            if user:
                mdp_temp = ''.join(random.choices(string.ascii_letters + string.digits, k=12))
                try:
                    send_mail(
                        subject='Réinitialisation de votre mot de passe CB Bretagne',
                        message=(
                            f'Bonjour {user.first_name or user.username},\n\n'
                            f'Votre mot de passe a été réinitialisé suite à votre demande.\n\n'
                            f'Identifiant : {user.username}\n'
                            f'Mot de passe temporaire : {mdp_temp}\n\n'
                            f'Connectez-vous ici : {settings.SITE_URL}/login/\n\n'
                            f'Pensez à changer votre mot de passe après connexion.\n\n'
                            f'CB Bretagne'
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[email],
                        fail_silently=False,
                    )
                    # MDP changé seulement si l'email est parti — évite de bloquer l'accès
                    user.set_password(mdp_temp)
                    user.save()
                except Exception as e:
                    logger.error('Password reset email error (user %s): %s', user.username, e)
            # Même message qu'un email inexistant pour éviter l'énumération
            succes = True

    return render(request, 'core/mot_de_passe_oublie.html', {
        'succes': succes,
        'erreur': erreur,
        'domaine': DOMAINE_AUTORISE,
    })


@login_required
def profil_view(request):
    profil = get_profil(request.user)
    if request.method == 'POST':
        # Infos personnelles
        request.user.first_name = request.POST.get('first_name', '').strip()
        request.user.last_name = request.POST.get('last_name', '').strip()
        request.user.email = request.POST.get('email', '').strip()
        request.user.save()
        # Préférences
        taux = request.POST.get('taux_mo_defaut', '').strip().replace(',', '.')
        if taux:
            try:
                profil.taux_mo_defaut = Decimal(taux)
            except Exception:
                messages.error(request, 'Taux MO invalide.')
                return redirect('core:profil')
        profil.saisie_ht = request.POST.get('saisie_ht') == 'on'
        profil.conditions_devis = request.POST.get('conditions_devis', '').strip()
        profil.conditions_facture = request.POST.get('conditions_facture', '').strip()
        profil.coordonnees_cb = request.POST.get('coordonnees_cb', '').strip()
        couleur = request.POST.get('couleur', '')
        if couleur in PALETTE_COULEURS:
            profil.couleur = couleur
        profil.save()
        messages.success(request, 'Profil mis à jour.')
        return redirect('core:profil')
    return render(request, 'core/profil.html', {'palette_couleurs': PALETTE_COULEURS})

    
# ══════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════

def _build_period_presets(today):
    """Construit les presets de période pour la barre de filtres production."""
    import calendar as _cal
    mois_fr = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
               'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
    presets = []
    # 6 derniers mois (mois courant inclus)
    y, m = today.year, today.month
    for _ in range(6):
        d1 = date(y, m, 1)
        d2 = date(y, m, _cal.monthrange(y, m)[1])
        label = f'{mois_fr[m]} {y}' if y == today.year else f'{mois_fr[m]} {y}'
        presets.append({'label': label, 'debut': d1.isoformat(), 'fin': d2.isoformat()})
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    # Trimestres de l'année courante
    yr = today.year
    for q, (qm1, qm2) in enumerate([(1,3),(4,6),(7,9),(10,12)], start=1):
        d1 = date(yr, qm1, 1)
        d2 = date(yr, qm2, _cal.monthrange(yr, qm2)[1])
        if d1 <= today:
            presets.append({'label': f'T{q} {yr}', 'debut': d1.isoformat(), 'fin': d2.isoformat()})
    # Année courante
    presets.append({'label': str(yr), 'debut': f'{yr}-01-01', 'fin': f'{yr}-12-31'})
    return presets


@login_required
def dashboard(request):
    profil = get_profil(request.user)
    visibles, disponibles = resolve_dashboard(profil, request.user)
    return render(request, 'core/dashboard.html', {
        'widgets': visibles,
        'widgets_disponibles': disponibles,
    })


@login_required
@require_POST
def dashboard_save(request):
    """Enregistre la disposition du tableau de bord (ordre + hidden + scope).

    Ignore les ids inconnus et les widgets compta si l'utilisateur n'y a pas
    droit (cf. sanitize_config). Renvoie du JSON.
    """
    profil = get_profil(request.user)
    payload, err = parse_json_request(request)
    if err:
        return err

    widgets = sanitize_config(payload.get('widgets'), request.user)
    profil.dashboard_config = {'widgets': widgets}
    profil.save(update_fields=['dashboard_config'])
    return JsonResponse({'ok': True})


# ══════════════════════════════════════════
#  CLIENTS
# ══════════════════════════════════════════

@login_required
def clients_list(request):
    nom = request.GET.get('nom', '').strip()
    code_postal = request.GET.get('code_postal', '').strip()
    departement = request.GET.get('departement', '').strip()
    ville = request.GET.get('ville', '').strip()
    portee = request.GET.get('portee', 'tous')
    type_client = request.GET.get('type_client', '').strip()

    clients = Client.objects.select_related('created_by__profil')
    if nom:
        clients = clients.filter(nom__icontains=nom)
    if code_postal:
        clients = clients.filter(code_postal__startswith=code_postal)
    if departement:
        clients = clients.filter(code_postal__startswith=departement)
    if ville:
        clients = clients.filter(ville__icontains=ville)
    if type_client:
        clients = clients.filter(type_client=type_client)
    if portee == 'moi':
        clients = clients.filter(created_by=request.user)
    elif portee == 'equipe':
        clients = clients.filter(created_by__in=get_collegues_ids(request.user))

    return render(request, 'core/clients.html', {
        'clients': clients.prefetch_related('contacts'),
        'is_admin': is_admin(request.user),
        'type_choices': Client.TYPE_CLIENT_CHOICES,
        'f_nom': nom,
        'f_code_postal': code_postal,
        'f_departement': departement,
        'f_ville': ville,
        'f_portee': portee,
        'f_type_client': type_client,
    })


@login_required
def client_search(request):
    """Recherche client (autocomplétion + filtre du panneau). Renvoie du JSON."""
    q = request.GET.get('q', '').strip()
    clients = Client.objects.all()
    if q:
        clients = clients.filter(nom__icontains=q)
    results = [
        {
            'id': c.pk,
            'nom': c.nom,
            'ville': c.ville,
            'code_postal': c.code_postal,
        }
        for c in clients[:20]
    ]
    return JsonResponse({'results': results})


@login_required
@require_POST
def client_quick_create(request):
    """Création rapide depuis l'écran de devis. Renvoie le client créé en JSON."""
    nom = request.POST.get('nom', '').strip()
    if not nom:
        return json_error('Le nom est obligatoire.')
    client = Client.objects.create(
        nom=nom,
        contact=request.POST.get('contact', ''),
        email=request.POST.get('email', ''),
        telephone=request.POST.get('telephone', ''),
        adresse=request.POST.get('adresse', ''),
        code_postal=request.POST.get('code_postal', ''),
        ville=request.POST.get('ville', ''),
        created_by=request.user,
    )
    return JsonResponse({'id': client.pk, 'nom': client.nom})


@login_required
@require_POST
def client_create(request):
    nom = request.POST.get('nom', '').strip()
    if not nom:
        messages.error(request, 'Le nom est obligatoire.')
        return redirect('core:clients')
    Client.objects.create(
        nom=nom,
        type_client=request.POST.get('type_client', 'particulier'),
        contact=request.POST.get('contact', ''),
        email=request.POST.get('email', ''),
        telephone=request.POST.get('telephone', ''),
        adresse=request.POST.get('adresse', ''),
        code_postal=request.POST.get('code_postal', ''),
        ville=request.POST.get('ville', ''),
        created_by=request.user,
    )
    messages.success(request, f'Client "{nom}" créé.')
    return redirect('core:clients')


@login_required
@require_POST
def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)
    nom = request.POST.get('nom', '').strip()
    if not nom:
        messages.error(request, 'Le nom est obligatoire.')
        return redirect('core:clients')
    client.nom = nom
    client.type_client = request.POST.get('type_client', client.type_client)
    client.contact = request.POST.get('contact', '')
    client.email = request.POST.get('email', '')
    client.telephone = request.POST.get('telephone', '')
    client.adresse = request.POST.get('adresse', '')
    client.code_postal = request.POST.get('code_postal', '')
    client.ville = request.POST.get('ville', '')
    client.save()
    messages.success(request, f'Client "{nom}" modifié.')
    return redirect('core:clients')


@login_required
@require_POST
def client_delete(request, pk):
    if not is_admin(request.user):
        messages.error(request, 'Action réservée à l\'administrateur.')
        return redirect('core:clients')
    client = get_object_or_404(Client, pk=pk)
    nom = client.nom
    client.delete()
    messages.success(request, f'Client "{nom}" supprimé.')
    return redirect('core:clients')

# ══════════════════════════════════════════
#  BIBLIOTHÈQUE PERSONNELLE (API JSON)
# ══════════════════════════════════════════

@login_required
def bibliotheque(request):
    profil = get_profil(request.user)
    return render(request, 'core/bibliotheque.html', {
        'taux_mo_js': str(profil.taux_mo_defaut).replace(',', '.'),
    })

@login_required
def biblio_api_get(request):
    biblio, _ = Bibliotheque.objects.get_or_create(user=request.user)
    profil = get_profil(request.user)
    return JsonResponse({
        'lignes': biblio.lignes,
        'taux_mo': float(profil.taux_mo_defaut),
    })

@login_required
@require_POST
def biblio_api_save(request):
    biblio, _ = Bibliotheque.objects.get_or_create(user=request.user)
    data, err = parse_json_request(request)
    if err:
        return err
    biblio.lignes = data.get('lignes', [])
    biblio.save()
    return JsonResponse({'ok': True})


# ══════════════════════════════════════════
#  BIBLIOTHÈQUE AIDES (partagée)
# ══════════════════════════════════════════

@login_required
def aides_page(request):
    return render(request, 'core/aides.html', {})


@login_required
def aides_api_get(request):
    aides = BibliothequeAides.objects.select_related('created_by').all()
    return JsonResponse({
        'aides': [
            {
                'id': a.pk,
                'description': a.description,
                'type_ligne': a.type_ligne,
                'montant_defaut': float(a.montant_defaut) if a.montant_defaut is not None else None,
                'unite': a.unite,
                'organisme': a.organisme,
            }
            for a in aides
        ]
    })


@login_required
@require_POST
def aides_api_save(request):
    data, err = parse_json_request(request)
    if err:
        return err
    description = data.get('description', '').strip()
    if not description:
        return json_error('Description requise')
    type_ligne = data.get('type_ligne', 'FIN')
    if type_ligne not in ('FMO', 'FMAT', 'FIN', 'FINX'):
        return json_error('Type invalide')
    montant_raw = data.get('montant_defaut')
    montant = to_decimal(montant_raw)
    aide = BibliothequeAides.objects.create(
        description=description,
        type_ligne=type_ligne,
        montant_defaut=montant,
        unite=data.get('unite', 'forfait'),
        organisme=data.get('organisme', ''),
        created_by=request.user,
    )
    return JsonResponse({
        'ok': True,
        'aide': {
            'id': aide.pk,
            'description': aide.description,
            'type_ligne': aide.type_ligne,
            'montant_defaut': float(aide.montant_defaut) if aide.montant_defaut is not None else None,
            'unite': aide.unite,
            'organisme': aide.organisme,
        }
    })


@login_required
@require_POST
def aide_delete(request, pk):
    aide = get_object_or_404(BibliothequeAides, pk=pk)
    aide.delete()
    return JsonResponse({'ok': True})


# ══════════════════════════════════════════
#  DEVIS — LISTE
# ══════════════════════════════════════════

@login_required
def devis_list(request):
    qs = Devis.objects.select_related(
        'client', 'equipe__service__territoire', 'created_by__profil'
    )

    # Filtres
    status = request.GET.get('status', '')
    client_id = request.GET.get('client', '')
    equipe_id = request.GET.get('equipe', '')
    service_id = request.GET.get('service', '')
    territoire_id = request.GET.get('territoire', '')
    auteur_id = request.GET.get('auteur', '')
    q = request.GET.get('q', '')

    if status:
        qs = qs.filter(status=status)
    if client_id:
        qs = qs.filter(client_id=client_id)
    if equipe_id:
        qs = qs.filter(equipe_id=equipe_id)
    if service_id:
        qs = qs.filter(equipe__service_id=service_id)
    if territoire_id:
        qs = qs.filter(equipe__service__territoire_id=territoire_id)
    if auteur_id:
        qs = qs.filter(created_by_id=auteur_id)
    if q:
        qs = qs.filter(
            Q(chantier__icontains=q)
            | Q(client__nom__icontains=q)
            | Q(reference__icontains=q)
        )

    # Prefetch posé sur le queryset final (après tous les filtres) pour que
    # `attacher_totaux_devis` calcule les totaux en mémoire, sans N+1.
    qs = qs.prefetch_related('lignes', 'factures')

    page_obj, base_qs = paginer(request, qs)
    # Totaux calculés en mémoire sur la page courante uniquement (pas de N+1).
    attacher_totaux_devis(page_obj.object_list)

    return render(request, 'core/devis_list.html', {
        'devis': page_obj,
        'page_obj': page_obj,
        'base_qs': base_qs,
        'clients': Client.objects.all(),
        'equipes': Equipe.objects.select_related('service__territoire').all(),
        'services': Service.objects.select_related('territoire').all(),
        'territoires': Territoire.objects.all(),
        'auteurs': User.objects.filter(devis_crees__isnull=False).distinct().order_by('first_name', 'username'),
        'status_filter': status,
        'client_filter': client_id,
        'equipe_filter': equipe_id,
        'service_filter': service_id,
        'territoire_filter': territoire_id,
        'auteur_filter': auteur_id,
        'q': q,
        'peut_supprimer': {d.pk: peut_supprimer_devis(request.user, d) for d in page_obj},
    })


# ══════════════════════════════════════════
#  DEVIS — CRÉATION
# ══════════════════════════════════════════

@login_required
def devis_create(request):
    profil = get_profil(request.user)
    if request.method == 'POST':
        client_id = request.POST.get('client')
        chantier = request.POST.get('chantier', '').strip()
        if not client_id or not chantier:
            messages.error(request, 'Client et chantier sont obligatoires.')
            return redirect('core:devis-list')
        client = get_object_or_404(Client, pk=client_id)
        equipe_id = request.POST.get('equipe') or None
        try:
            validite_jours = int(request.POST.get('validite_jours') or 30)
        except (TypeError, ValueError):
            validite_jours = 30
        try:
            taux_mo = Decimal(str(request.POST.get('taux_mo') or profil.taux_mo_defaut))
        except (InvalidOperation, TypeError, ValueError):
            taux_mo = profil.taux_mo_defaut
        devis = Devis.objects.create(
            reference=gen_numero_sequence('DEV', Devis, 'reference'),
            client=client,
            chantier=chantier,
            equipe_id=equipe_id,
            date_validite=timezone.localdate() + timedelta(days=validite_jours),
            taux_mo=taux_mo,
            notes=request.POST.get('notes', ''),
            conditions_devis=profil.conditions_devis,
            created_by=request.user,
        )
        add_audit(
            request.user,
            f"Création devis {devis.reference} — {client.nom} · {chantier}",
            devis=devis
        )
        return redirect('core:devis-detail', pk=devis.pk)

    # Déterminer les équipes disponibles selon le profil
    if profil.service:
        equipes = Equipe.objects.filter(
            service__territoire=profil.get_territoire()
        ).select_related('service')
    else:
        equipes = Equipe.objects.select_related('service').all()

    return render(request, 'core/devis_form.html', {
        'clients': Client.objects.all(),
        'equipes': equipes,
    })


# ══════════════════════════════════════════
#  DEVIS — DÉTAIL
# ══════════════════════════════════════════

@login_required
def devis_detail(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_voir_devis(request.user, devis):
        messages.error(request, "Vous n'avez pas accès à ce devis.")
        return redirect('core:devis-list')
    profil = get_profil(request.user)
    factures = devis.factures.all()
    audit_logs = devis.audit_logs.all()
    circuit_steps = [
        ('ti-file-plus', 'var(--gray-lt)', 'var(--gray-bd)', 'Brouillon auto'),
        ('ti-user-check', '#E6F1FB', '#B5D4F4', 'Validation comptable'),
        ('ti-circle-check', 'var(--teal-lt)', 'var(--teal-bd)', 'Validée'),
        ('ti-send', 'var(--prune-lt)', 'var(--prune-bd)', 'Envoyée'),
        ('ti-cash', '#EAF3DE', '#C0DD97', 'Payée'),
    ]
    taux_mo_js = str(devis.taux_mo).replace(',', '.')
    articles_biblio = []
    return render(request, 'core/devis_detail.html', {
        'devis': devis,
        'factures': factures,
        'audit_logs': audit_logs,
        'articles_biblio': articles_biblio,
        'circuit_steps': circuit_steps,
        'taux_mo_js': taux_mo_js,
        'saisie_ht': profil.saisie_ht,  # ← ajouté
        'peut_modifier': peut_modifier_devis(request.user, devis),  # éditeur verrouillé si hors équipe
        'clients': Client.objects.all(),
        'equipes': Equipe.objects.select_related('service__territoire').all(),
    })


@login_required
@require_POST
def devis_status(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_modifier_devis(request.user, devis):
        messages.error(request, 'Vous ne pouvez pas modifier ce devis.')
        return redirect('core:devis-detail', pk=pk)
    old = devis.status
    devis.status = request.POST.get('status', devis.status)
    devis.save()
    add_audit(
        request.user,
        f"Statut devis {devis.reference} : {old} → {devis.status}",
        devis=devis
    )
    return redirect('core:devis-detail', pk=pk)


@login_required
def devis_duplicate(request, pk):
    src = get_object_or_404(Devis, pk=pk)
    if not peut_voir_devis(request.user, src):
        messages.error(request, "Vous n'avez pas accès à ce devis.")
        return redirect('core:devis-list')
    profil = get_profil(request.user)
    new_devis = Devis.objects.create(
        reference=gen_numero_sequence('DEV', Devis, 'reference'),
        client=src.client,
        chantier=src.chantier + ' (copie)',
        equipe=src.equipe,
        status='draft',
        date_validite=timezone.localdate() + timedelta(days=30),
        taux_mo=src.taux_mo,
        notes=src.notes,
        conditions_devis=src.conditions_devis,
        fin_group_title=src.fin_group_title,
        zone_financement_ext=src.zone_financement_ext,
        created_by=request.user,
    )

    def copy_lignes(lignes, parent=None):
        for l in lignes:
            new_l = LigneDevis.objects.create(
                devis=new_devis, parent=parent,
                type_ligne=l.type_ligne,
                description=l.description,
                quantite=l.quantite,
                unite=l.unite,
                cout_unitaire=l.cout_unitaire,
                ordre=l.ordre,
            )
            copy_lignes(l.enfants.all(), parent=new_l)

    copy_lignes(src.lignes.filter(parent=None))
    add_audit(
        request.user,
        f"Duplication {src.reference} → {new_devis.reference}",
        devis=new_devis
    )
    return redirect('core:devis-detail', pk=new_devis.pk)


@login_required
@require_POST
def devis_delete(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_supprimer_devis(request.user, devis):
        messages.error(request, 'Vous ne pouvez pas supprimer ce devis.')
        return redirect('core:devis-list')
    ref = devis.reference
    devis.delete()
    messages.success(request, f'Devis {ref} supprimé.')
    return redirect('core:devis-list')


def assign_numbers_python(lignes, prefix=''):
    flat = []
    local_counter = [0]

    for ligne in lignes:
        if ligne.type_ligne in ('TITRE', 'S'):
            local_counter[0] += 1
            num = f"{prefix}{local_counter[0]}" if prefix else str(local_counter[0])
        else:
            num = ''

        ligne.num = num
        flat.append(ligne)

        # On descend uniquement dans les TITRE
        if ligne.type_ligne == 'TITRE':
            enfants = list(ligne.enfants.all())
            if enfants:
                flat.extend(assign_numbers_python(enfants, prefix=f"{num}."))

    return flat


def _render_pdf(request, template_name, context, filename):
    from weasyprint import HTML
    from django.template.loader import render_to_string
    html = render_to_string(template_name, context, request=request)
    pdf = HTML(string=html, base_url=request.build_absolute_uri('/')).write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    safe_name = re.sub(r'[^\w\-.]', '_', filename)
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response


@login_required
def devis_pdf(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_voir_devis(request.user, devis):
        messages.error(request, "Vous n'avez pas accès à ce devis.")
        return redirect('core:devis-list')
    factures = devis.factures.exclude(status='cancelled')
    params = ParametresAssociation.get()

    racines_pos = list(
        devis.lignes.filter(parent=None).exclude(type_ligne='FIN').prefetch_related('enfants')
    )
    racines_fin = list(
        devis.lignes.filter(parent=None, type_ligne='FIN').prefetch_related('enfants')
    )
    lignes_pos = assign_numbers_python(racines_pos)
    lignes_fin = assign_numbers_python(racines_fin)
    expiry = devis.date_validite

    ctx = {
        'devis': devis,
        'factures': factures,
        'lignes_pos': lignes_pos,
        'lignes_fin': lignes_fin,
        'params': params,
        'expiry': expiry,
        'coordonnees_cb': devis.coordonnees_cb,
    }
    if request.GET.get('download') == '1':
        return _render_pdf(request, 'core/devis_pdf.html', ctx, f'{devis.reference}.pdf')
    return render(request, 'core/devis_pdf.html', ctx)

@login_required
def devis_export_excel(request, pk):
    import html as html_module
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    def strip_html(text):
        if not text or '<' not in text:
            return text or ''
        # <div> → saut de ligne (sauf le premier)
        t = re.sub(r'<div>', '\n', text, flags=re.IGNORECASE)
        # <br> variants → saut de ligne
        t = re.sub(r'<br\s*/?>', '\n', t, flags=re.IGNORECASE)
        # Supprimer toutes les balises restantes
        t = re.sub(r'<[^>]+>', '', t)
        # Décoder les entités HTML (&amp; &lt; etc.)
        t = html_module.unescape(t)
        # Nettoyer les lignes vides et espaces superflus
        lines = [l.strip() for l in t.splitlines() if l.strip()]
        return '\n'.join(lines)

    devis = get_object_or_404(Devis, pk=pk)
    if not peut_voir_devis(request.user, devis):
        messages.error(request, "Vous n'avez pas accès à ce devis.")
        return redirect('core:devis-list')

    PRUNE    = '67123A'
    PRUNE_LT = 'F2DCDB'
    GRAY     = 'E0E0E0'
    AMBER_LT = 'FFF2CC'
    GREEN_LT = 'E2EFDA'
    GREEN_MID = 'C6EFCE'
    GREEN_DK  = '375623'

    def fill(hex_color):
        return PatternFill('solid', fgColor=hex_color)

    FN = Font(name='Calibri', size=10)
    FB = Font(name='Calibri', size=10, bold=True)
    FW = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    FWL = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    FT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    FC = Font(name='Calibri', size=10, bold=True, color=PRUNE)

    AL = Alignment(vertical='center', wrap_text=True)
    AR = Alignment(horizontal='right', vertical='center')
    AC = Alignment(horizontal='center', vertical='center')
    AT = Alignment(wrap_text=True, vertical='top')

    FMT_EUR = '#,##0.00 "€"'
    FMT_QTY = '#,##0.###'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Devis'

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    row = 1

    def c(r, col, value='', font=None, fill_=None, align=None, fmt=None):
        obj = ws.cell(row=r, column=col, value=value)
        if font:  obj.font      = font
        if fill_: obj.fill      = fill_
        if align: obj.alignment = align
        if fmt:   obj.number_format = fmt
        return obj

    def full_row(r, font=None, fill_=None, value=None, align=None, height=16):
        ws.row_dimensions[r].height = height
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        for col in range(1, 7):
            ws.cell(r, col).fill = fill_ or PatternFill()
        if value is not None:
            ws.cell(r, 1).value = value
        if font:  ws.cell(r, 1).font = font
        if align: ws.cell(r, 1).alignment = align

    def total_row(r, label, amount, font_label=None, font_amount=None, fill_=None, height=16):
        ws.row_dimensions[r].height = height
        for col in range(1, 7):
            ws.cell(r, col).fill = fill_ or PatternFill()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = font_label or FB
        ws.cell(r, 1).alignment = AR
        ws.cell(r, 6).value = float(amount)
        ws.cell(r, 6).font = font_amount or FB
        ws.cell(r, 6).number_format = FMT_EUR
        ws.cell(r, 6).alignment = AR

    # ── En-tête ──────────────────────────────────────────────────
    ws.row_dimensions[row].height = 28
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    for col in range(1, 7):
        ws.cell(row, col).fill = fill(PRUNE)
    ws.cell(row, 1).value = f"DEVIS TRAVAUX — {devis.reference}"
    ws.cell(row, 1).font = FT
    ws.cell(row, 1).alignment = AL
    ws.cell(row, 5).value = f"Date : {devis.date_creation.strftime('%d/%m/%Y')}"
    ws.cell(row, 5).font = Font(name='Calibri', size=10, color='FFFFFF')
    ws.cell(row, 5).alignment = AR
    row += 1

    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    for col in range(1, 7):
        ws.cell(row, col).fill = fill(PRUNE_LT)
    ws.cell(row, 1).value = (
        f"Statut : {devis.get_status_display()}"
        f"   |   Chantier : {devis.chantier}"
    )
    ws.cell(row, 1).font = FB
    ws.cell(row, 1).alignment = AL
    row += 1

    def info_row(r, label, value):
        ws.row_dimensions[r].height = 14
        ws.cell(r, 1).value = label
        ws.cell(r, 1).font = FB
        ws.cell(r, 2).value = value
        ws.cell(r, 2).font = FN
        ws.cell(r, 2).alignment = AL

    info_row(row, 'Client :', str(devis.client))
    if devis.client.telephone:
        ws.cell(row, 4).value = f"Tél : {devis.client.telephone}"
        ws.cell(row, 4).font = FN
    if devis.client.email:
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row, 5).value = devis.client.email
        ws.cell(row, 5).font = FN
    row += 1

    adresse_client = ' — '.join(filter(None, [
        devis.client.adresse,
        f"{devis.client.code_postal} {devis.client.ville}".strip() or None,
    ]))
    if adresse_client:
        info_row(row, '', adresse_client)
        row += 1

    chantier_adresse = ' — '.join(filter(None, [
        devis.chantier_adresse1,
        devis.chantier_adresse2,
        f"{devis.chantier_cp} {devis.chantier_ville}".strip() or None,
    ]))
    if chantier_adresse:
        info_row(row, 'Adresse chantier :', chantier_adresse)
        row += 1

    if devis.equipe:
        info_row(row, 'Équipe :', str(devis.equipe))
        row += 1

    if devis.coordonnees_cb:
        info_row(row, 'Contact CB :', devis.coordonnees_cb.replace('\n', ' | '))
        row += 1

    if devis.date_validite:
        info_row(row, "Valable jusqu'au :", devis.date_validite.strftime('%d/%m/%Y'))
        row += 1

    row += 1  # séparateur

    # En-tête colonnes
    ws.row_dimensions[row].height = 18
    for col, label in enumerate(['N°', 'Description', 'Unité', 'Quantité', 'P.U. HT', 'Total HT'], 1):
        ws.cell(row, col).value = label
        ws.cell(row, col).font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        ws.cell(row, col).fill = fill('444444')
        ws.cell(row, col).alignment = AC
    row += 1

    # ── Lignes de devis ──────────────────────────────────────────
    racines_pos = list(
        devis.lignes.filter(parent=None).exclude(type_ligne='FIN')
        .prefetch_related('enfants', 'enfants__enfants', 'enfants__enfants__enfants')
        .order_by('ordre')
    )
    racines_fin = list(
        devis.lignes.filter(parent=None, type_ligne='FIN').order_by('ordre')
    )

    titres      = [l for l in racines_pos if l.type_ligne == 'TITRE']
    autres_rac  = [l for l in racines_pos if l.type_ligne != 'TITRE']

    def write_enfants(r, parent_ligne, depth=1):
        for enfant in parent_ligne.enfants.all():
            r = write_ligne(r, enfant, depth)
        return r

    def write_ligne(r, ligne, depth=0):
        ws.row_dimensions[r].height = 15
        indent = '  ' * depth
        t = ligne.type_ligne

        if t == 'TITRE':
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            for col in range(1, 7):
                ws.cell(r, col).fill = fill(PRUNE_LT)
            ws.cell(r, 1).value = indent + strip_html(ligne.description)
            ws.cell(r, 1).font = FC
            ws.cell(r, 1).alignment = AL
            ws.cell(r, 6).value = float(ligne.total())
            ws.cell(r, 6).font = FC
            ws.cell(r, 6).number_format = FMT_EUR
            ws.cell(r, 6).alignment = AR
            r += 1
            r = write_enfants(r, ligne, depth + 1)
            # SS TOTAL du titre
            ws.row_dimensions[r].height = 15
            for col in range(1, 7):
                ws.cell(r, col).fill = fill(GRAY)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            ws.cell(r, 1).value = f"SS TOTAL — {strip_html(ligne.description)}"
            ws.cell(r, 1).font = FB
            ws.cell(r, 1).alignment = AR
            ws.cell(r, 6).value = float(ligne.total())
            ws.cell(r, 6).font = FB
            ws.cell(r, 6).number_format = FMT_EUR
            ws.cell(r, 6).alignment = AR
            r += 1
            ws.row_dimensions[r].height = 6   # ligne vide entre LOTs
            r += 1
            return r

        # Ligne ordinaire
        num = getattr(ligne, 'num', '')
        ws.cell(r, 1).value = indent + (num or '')
        ws.cell(r, 1).font = FN
        ws.cell(r, 1).alignment = AC

        ws.cell(r, 2).value = indent + strip_html(ligne.description)
        ws.cell(r, 2).font = FN
        ws.cell(r, 2).alignment = AL

        ws.cell(r, 3).value = ligne.unite or ''
        ws.cell(r, 3).font = FN
        ws.cell(r, 3).alignment = AC

        if t not in ('FMO', 'FMAT'):
            ws.cell(r, 4).value = float(ligne.quantite)
            ws.cell(r, 4).number_format = FMT_QTY
            ws.cell(r, 4).font = FN
            ws.cell(r, 4).alignment = AR

        if ligne.cout_unitaire is not None:
            ws.cell(r, 5).value = float(ligne.cout_unitaire)
            ws.cell(r, 5).number_format = FMT_EUR
            ws.cell(r, 5).font = FN
            ws.cell(r, 5).alignment = AR

        total_val = float(ligne.total())
        ws.cell(r, 6).value = total_val if total_val != 0 else None
        ws.cell(r, 6).number_format = FMT_EUR
        ws.cell(r, 6).font = FB if t in ('FMO', 'FMAT') else FN
        ws.cell(r, 6).alignment = AR

        if t in ('FMO', 'FMAT'):
            for col in range(1, 7):
                ws.cell(r, col).fill = fill(AMBER_LT)

        r += 1
        r = write_enfants(r, ligne, depth + 1)
        return r

    # Section Travaux réalisés
    full_row(row, font=FW, fill_=fill(PRUNE),
             value='TRAVAUX RÉALISÉS', align=AL, height=18)
    for col in range(2, 7):
        ws.cell(row, col).fill = fill(PRUNE)
    row += 1

    for titre in titres:
        row = write_ligne(row, titre, depth=0)

    for ligne in autres_rac:
        row = write_ligne(row, ligne, depth=0)

    # SS TOTAL Travaux
    total_travaux = sum(l.total() for l in titres + autres_rac)
    total_row(row, 'SS TOTAL TRAVAUX', total_travaux, fill_=fill(GRAY), height=16)
    row += 1

    # Total brut net de taxes
    row += 1
    total_row(row, 'TOTAL COÛT DES TRAVAUX NET DE TAXES',
              devis.total_brut(),
              font_label=FW, font_amount=FWL, fill_=fill(PRUNE), height=20)
    row += 1

    # Section Aides / Financements
    if racines_fin:
        row += 1
        full_row(row, font=FW, fill_=fill(GREEN_DK),
                 value='AIDES NOTIFIÉES', align=AL, height=18)
        for col in range(2, 7):
            ws.cell(row, col).fill = fill(GREEN_DK)
        row += 1
        for ligne in racines_fin:
            ws.row_dimensions[row].height = 15
            for col in range(1, 7):
                ws.cell(row, col).fill = fill(GREEN_LT)
            ws.cell(row, 2).value = strip_html(ligne.description)
            ws.cell(row, 2).font = FN
            ws.cell(row, 2).alignment = AL
            ws.cell(row, 3).value = ligne.unite or ''
            ws.cell(row, 3).font = FN
            ws.cell(row, 3).alignment = AC
            ws.cell(row, 6).value = float(ligne.total())
            ws.cell(row, 6).number_format = FMT_EUR
            ws.cell(row, 6).font = FN
            ws.cell(row, 6).alignment = AR
            row += 1
        total_row(row, 'SS TOTAL Aides', devis.total_financement(),
                  fill_=fill(GREEN_MID), height=16)
        row += 1

    # Montant dû
    row += 1
    total_row(row, "MONTANT DÛ PAR LE MAÎTRE D'OUVRAGE",
              devis.net_client(),
              font_label=FW, font_amount=FWL, fill_=fill(PRUNE), height=20)
    row += 1

    # Acomptes versés
    acomptes = list(devis.factures.filter(type_doc='acompte', status='paid'))
    if acomptes:
        row += 1
        ws.row_dimensions[row].height = 14
        ws.cell(row, 1).value = 'NB : mettre les dates de versement'
        ws.cell(row, 1).font = Font(name='Calibri', size=9, italic=True)
        ws.cell(row, 2).value = 'Acompte déjà versé'
        ws.cell(row, 2).font = FB
        row += 1
        for f in acomptes:
            ws.row_dimensions[row].height = 14
            label = f.libelle or f.get_reference()
            if f.date_versement:
                label += f" (versé le {f.date_versement.strftime('%d/%m/%Y')})"
            ws.cell(row, 2).value = label
            ws.cell(row, 2).font = FN
            ws.cell(row, 6).value = float(f.montant)
            ws.cell(row, 6).number_format = FMT_EUR
            ws.cell(row, 6).alignment = AR
            row += 1
        total_acomptes = sum(f.montant for f in acomptes)
        total_row(row, 'Reste à payer',
                  devis.net_client() - total_acomptes,
                  fill_=fill(GRAY), height=16)
        row += 1

    # Conditions
    if devis.conditions_devis:
        row += 1
        ws.row_dimensions[row].height = 14
        ws.cell(row, 1).value = 'Conditions :'
        ws.cell(row, 1).font = FB
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).value = devis.conditions_devis
        ws.cell(row, 1).font = Font(name='Calibri', size=9)
        ws.cell(row, 1).alignment = AT
        ws.row_dimensions[row].height = max(
            30, min(len(devis.conditions_devis) // 5, 100)
        )
        row += 2

    # Signatures
    ws.row_dimensions[row].height = 14
    ws.cell(row, 2).value = "Pour le Maître d'ouvrage"
    ws.cell(row, 2).font = FB
    ws.cell(row, 5).value = "Pour les Compagnons Bâtisseurs Bretagne"
    ws.cell(row, 5).font = FB
    row += 1
    ws.cell(row, 2).value = 'Signature précédée de la mention "Bon pour accord"'
    ws.cell(row, 2).font = Font(name='Calibri', size=9, italic=True)
    ws.cell(row, 5).value = 'Signature + Cachet'
    ws.cell(row, 5).font = Font(name='Calibri', size=9, italic=True)

    # ── Réponse ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    ref_clean = devis.reference.replace('/', '-')
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="devis_{ref_clean}.xlsx"'
    return response


@login_required
@require_POST
def devis_entete_save(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_modifier_devis(request.user, devis):
        return json_error_permission()
    data, err = parse_json_request(request)
    if err:
        return err

    # Client
    client_id = data.get('client_id')
    if client_id:
        client = get_object_or_404(Client, pk=client_id)
        devis.client = client

    # Équipe (peut être vide)
    equipe_id = data.get('equipe_id')
    if equipe_id:
        devis.equipe = get_object_or_404(Equipe, pk=equipe_id)
    else:
        devis.equipe = None

    # Chantier (obligatoire)
    chantier = data.get('chantier', '').strip()
    if chantier:
        devis.chantier = chantier

    # Adresse chantier
    devis.chantier_adresse1 = data.get('chantier_adresse1', '').strip()
    devis.chantier_adresse2 = data.get('chantier_adresse2', '').strip()
    devis.chantier_cp       = data.get('chantier_cp', '').strip()
    devis.chantier_ville    = data.get('chantier_ville', '').strip()

    # Notes et conditions
    devis.notes            = data.get('notes', '').strip()
    devis.conditions_devis = data.get('conditions_devis', '').strip()
    devis.coordonnees_cb   = data.get('coordonnees_cb', '').strip()

    # Date de validité
    date_validite_str = data.get('date_validite', '')
    if date_validite_str:
        try:
            from datetime import datetime as dt
            devis.date_validite = dt.strptime(date_validite_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Taux MO
    taux_mo = data.get('taux_mo')
    if taux_mo is not None:
        try:
            devis.taux_mo = Decimal(str(taux_mo))
        except Exception:
            pass

    devis.save()
    add_audit(
        request.user,
        f"Modification en-tête devis {devis.reference}",
        devis=devis
    )
    return JsonResponse({
        'ok': True,
        'client_nom': devis.client.nom,
        'chantier': devis.chantier,
        'equipe': str(devis.equipe) if devis.equipe else '',
        'taux_mo': float(devis.taux_mo),
    })

    
# ══════════════════════════════════════════
#  LIGNES DEVIS (API JSON)
# ══════════════════════════════════════════

@login_required
def lignes_get(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_voir_devis(request.user, devis):
        return json_error_permission()
    racines = devis.lignes.filter(parent=None)
    data = [ligne_to_dict(l) for l in racines]
    return JsonResponse({
        'lignes': data,
        'taux_mo': float(devis.taux_mo),
        'fin_group_title': devis.fin_group_title or 'Financements',
        'zone_financement': devis.zone_financement,
        'zone_financement_ext': devis.zone_financement_ext,
    })


@login_required
@require_POST
def lignes_save(request, pk):
    devis = get_object_or_404(Devis, pk=pk)
    if not peut_modifier_devis(request.user, devis):
        return json_error_permission()
    data, err = parse_json_request(request)
    if err:
        return err
    lignes = data.get('lignes', [])
    fin_group_title = data.get('fin_group_title', 'Financements')
    zone_financement = data.get('zone_financement', False)
    zone_financement_ext = data.get('zone_financement_ext', False)

    devis.lignes.all().delete()

    create_lignes = build_lignes_creator(LigneDevis, {'devis': devis}, with_aide=True)
    create_lignes(lignes)
    devis.fin_group_title = fin_group_title
    devis.zone_financement = zone_financement
    devis.zone_financement_ext = zone_financement_ext
    devis.updated_at = timezone.now()
    devis.save()

    return JsonResponse({
        'ok': True,
        'total_brut': float(devis.total_brut()),
        'net_client': float(devis.net_client()),
    })


# ══════════════════════════════════════════
#  FACTURES
# ══════════════════════════════════════════

@login_required
def factures_list(request):
    # Factures de chantier uniquement (liées à un devis, hors avoirs).
    # Les factures compta (structure/appel) et les avoirs ont leurs propres listes.
    factures = Facture.objects.select_related(
        'devis', 'devis__client', 'devis__equipe', 'created_by__profil'
    ).prefetch_related('avoirs').filter(devis__isnull=False).exclude(type_doc='avoir')

    q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')
    auteur_id = request.GET.get('auteur', '')
    if q:
        factures = factures.filter(
            Q(numero__icontains=q) | Q(notes__icontains=q)
            | Q(devis__client__nom__icontains=q) | Q(destinataire__icontains=q)
        )
    if status_filter:
        factures = factures.filter(status=status_filter)
    if auteur_id:
        factures = factures.filter(created_by_id=auteur_id)

    page_obj, base_qs = paginer(request, factures)

    return render(request, 'core/factures_list.html', {
        'factures': page_obj,
        'page_obj': page_obj,
        'base_qs': base_qs,
        'auteurs': User.objects.filter(
            factures_creees__devis__isnull=False
        ).exclude(factures_creees__type_doc='avoir').distinct().order_by('first_name', 'username'),
        'auteur_filter': auteur_id,
        'status_filter': status_filter,
        'q': q,
    })


@login_required
def avoirs_list(request):
    """
    Liste de tous les avoirs (chantier + compta).
    Lecture partagée pour les avoirs de chantier ; les avoirs compta (sans devis)
    ne sont visibles que par les rôles compta — chaque ligne affichée est donc
    accessible à l'utilisateur.
    """
    avoirs = Facture.objects.filter(type_doc='avoir').select_related(
        'devis', 'client', 'facture_origine', 'created_by__profil'
    ).order_by('-created_at')
    if not peut_acceder_compta(request.user):
        avoirs = avoirs.filter(devis__isnull=False)

    q = request.GET.get('q', '').strip()
    auteur_id = request.GET.get('auteur', '')
    if q:
        avoirs = avoirs.filter(
            Q(numero__icontains=q) | Q(client__nom__icontains=q)
            | Q(facture_origine__numero__icontains=q) | Q(destinataire__icontains=q)
        )
    if auteur_id:
        avoirs = avoirs.filter(created_by_id=auteur_id)

    page_obj, base_qs = paginer(request, avoirs)
    return render(request, 'core/avoirs_list.html', {
        'avoirs': page_obj,
        'page_obj': page_obj,
        'base_qs': base_qs,
        'auteurs': User.objects.filter(factures_creees__type_doc='avoir').distinct().order_by('first_name', 'username'),
        'auteur_filter': auteur_id,
        'q': q,
    })


@login_required
def facture_create(request, devis_pk):
    devis = get_object_or_404(Devis, pk=devis_pk)
    if not peut_modifier_devis(request.user, devis):
        return HttpResponseForbidden()

    if request.method != 'POST':
        return redirect('core:devis-detail', devis_pk)

    type_doc     = request.POST.get('type_doc', 'facture')
    destinataire = request.POST.get('destinataire', '').strip()
    notes        = request.POST.get('notes', '').strip()

    if type_doc == 'acompte': 
        # -> facture acompte
        montant_raw = request.POST.get('montant', '').strip()
        try:
            montant = Decimal(montant_raw)
        except Exception:
            montant = Decimal('0')

        facture = Facture.objects.create(
            devis=devis,
            type_doc='acompte',
            destinataire=destinataire or str(devis.client),
            montant=montant,
            notes=notes,
            created_by=request.user,
        )
        AuditLog.objects.create(
            user=request.user,
            action=f"Facture d'acompte {facture.get_reference()} créée ({montant} €)",
            devis=devis,
            facture=facture,
        )
        return redirect('core:devis-detail', devis_pk)

    else:
        # -> facture normale
        try:
            echeance_jours = int(request.POST.get('echeance_jours') or 30)
        except (TypeError, ValueError):
            echeance_jours = 30
        date_echeance  = timezone.localdate() + timedelta(days=echeance_jours)

        facture = Facture.objects.create(
            devis=devis,
            type_doc='facture',
            destinataire=destinataire or str(devis.client),
            notes=notes,
            date_echeance=date_echeance,
            created_by=request.user,
        )
        # Copie des lignes du devis, quantités ajustées selon les factures précédentes
        deja_detail = calc_deja_par_source_detail(devis, facture)
        copier_lignes_devis_vers_facture(devis.lignes.filter(parent=None), facture,
                                          deja_par_source=deja_detail)

        AuditLog.objects.create(
            user=request.user,
            action=f"Facture {facture.get_reference()} créée",
            devis=devis,
            facture=facture,
        )
        return redirect('core:facture-detail', facture.pk)

@login_required
@require_POST
def facture_date_versement(request, pk):
    facture = get_object_or_404(Facture, pk=pk, type_doc='acompte')
    if not peut_modifier_facture(request.user, facture):
        return json_error_permission()
    data, err = parse_json_request(request)
    if err:
        return err
    date_str = data.get('date_versement', '').strip()
    try:
        facture.date_versement = dt.strptime(date_str, '%d/%m/%Y').date() if date_str else None
    except ValueError:
        return json_error('Format invalide (jj/mm/aaaa)')
    facture.save(update_fields=['date_versement'])
    return JsonResponse({'ok': True})

@login_required
@require_POST
def facture_valider(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_valider_facture(request.user, facture):
        messages.error(request, 'Validation réservée au comptable.')
        return redirect('core:devis-detail', pk=facture.devis.pk)
    if not facture.numero:
        facture.numero = gen_numero_facture(facture.type_doc)
    facture.status = 'validated'
    facture.validated_by = request.user
    facture.validated_at = timezone.now()
    facture.save()
    add_audit(
        request.user,
        f"Validation comptable {facture.numero} ({facture.destinataire})",
        devis=facture.devis, facture=facture
    )
    return redirect('core:devis-detail', pk=facture.devis.pk)


@login_required
@require_POST
def facture_status(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'
    if not peut_modifier_facture(request.user, facture):
        if is_ajax:
            return json_error_permission()
        messages.error(request, 'Vous ne pouvez pas modifier cette facture.')
        return redirect('core:devis-detail', pk=facture.devis.pk)
    old = facture.status
    new_status = request.POST.get('status', facture.status)
    facture.status = new_status
    facture.save()
    add_audit(
        request.user,
        f"Statut {facture.get_reference()} : {old} → {new_status}",
        devis=facture.devis, facture=facture
    )
    # Si appel AJAX (fetch depuis JS), retourner JSON
    if is_ajax:
        return JsonResponse({'ok': True})
    # Sinon redirect normal (formulaires HTML)
    return redirect('core:devis-detail', pk=facture.devis.pk)


@login_required
@require_POST
def facture_bypass(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_modifier_facture(request.user, facture):
        return json_error_permission()
    code = request.POST.get('code', '')
    stored = request.session.get(f'bypass_code_{pk}')
    if not stored or code != stored:
        return json_error('Code incorrect')
    if not facture.numero:
        facture.numero = gen_numero_facture(facture.type_doc)
    facture.status = 'validated'
    facture.bypass_validation = True
    facture.validated_by = request.user
    facture.validated_at = timezone.now()
    facture.save()
    del request.session[f'bypass_code_{pk}']
    add_audit(
        request.user,
        f"Bypass comptable {facture.numero} ({facture.destinataire}) — code e-mail utilisé",
        devis=facture.devis, facture=facture, bypass=True
    )
    return JsonResponse({'ok': True, 'redirect': f'/devis/{facture.devis.pk}/'})

@login_required
@require_POST
def facture_delete(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    # Vérification permission avant le statut (évite de donner une info sur l'existence)
    if not peut_supprimer_facture(request.user, facture):
        messages.error(request, 'Vous ne pouvez pas supprimer cette facture.')
        return redirect('core:devis-detail', pk=facture.devis.pk)
    if facture.status != 'draft':
        messages.error(request, 'Seules les factures en brouillon peuvent être supprimées.')
        return redirect('core:devis-detail', pk=facture.devis.pk)
    devis = facture.devis
    ref = facture.get_reference()
    add_audit(request.user, f"Suppression facture brouillon {ref}", devis=devis)
    facture.delete()
    messages.success(request, f'Facture {ref} supprimée.')
    return redirect('core:devis-detail', pk=devis.pk)

@login_required
def facture_bypass_send_code(request, pk):
    if not request.user.email:
        return json_error('Aucune adresse email sur votre compte. Contactez un administrateur.')

    facture = get_object_or_404(Facture, pk=pk)
    if not peut_modifier_facture(request.user, facture):
        return json_error_permission()

    code = ''.join(random.choices(string.digits, k=6))

    try:
        send_mail(
            subject=f'Code bypass — {facture.get_reference()}',
            message=(
                f'Bonjour {request.user.first_name or request.user.username},\n\n'
                f'Votre code de bypass pour la facture {facture.get_reference()} est :\n\n'
                f'    {code}\n\n'
                f'Ce code est valable 10 minutes.\n\n'
                f'CB Bretagne'
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[request.user.email],
            fail_silently=False,
        )
        request.session[f'bypass_code_{pk}'] = code
        return JsonResponse({'ok': True})
    except Exception as e:
        logger.error('Bypass email error (facture %s): %s', pk, e)
        return json_error("Impossible d'envoyer le code par email. Contactez un administrateur.")


def calc_deja_par_source_detail(devis, facture_courante):
    """
    Retourne {ligne_devis_id: {'montant': float, 'qty': Decimal, 'refs': {ref: montant}}}
    pour toutes les factures VALIDÉES du devis, en excluant la facture courante.

    Statuts comptabilisés : validated, sent, paid. Exclus : draft, cancelled, avoirs.
    Règle TITRE : si TITRE.quantite=0, on ne descend pas dans ses enfants (la section
    était exclue — ses lignes feuilles n'ont pas été facturées).
    """
    STATUTS_VALIDES = ('validated', 'sent', 'paid')
    deja = {}
    for f in devis.factures.filter(
        status__in=STATUTS_VALIDES, type_doc='facture'
    ).exclude(pk=facture_courante.pk):
        ref = f.get_reference()
        for lf in f.lignes.filter(parent=None):
            _agreger_deja(lf, deja, ref)
    return deja


def _agreger_deja(ligne, deja, ref_facture, titre_factor=Decimal('1')):
    """Accumule montant + qty + refs par source_id en tenant compte du facteur TITRE parent."""
    if ligne.type_ligne == 'TITRE' and ligne.quantite == 0:
        return  # section exclue — ne pas descendre

    if ligne.ligne_devis_source_id:
        sid = ligne.ligne_devis_source_id
        effective_total = float(ligne.total()) * float(titre_factor)
        effective_qty   = ligne.quantite * titre_factor
        e = deja.setdefault(sid, {'montant': 0.0, 'qty': Decimal('0'), 'refs': {}})
        e['montant'] += effective_total
        e['qty']     += effective_qty
        if effective_qty > 0:
            e['refs'][ref_facture] = e['refs'].get(ref_facture, 0.0) + effective_total

    child_factor = (titre_factor * ligne.quantite) if ligne.type_ligne == 'TITRE' else titre_factor
    for enfant in ligne.enfants.all():
        _agreger_deja(enfant, deja, ref_facture, child_factor)


# ──────────────────────────────────────────
#  VUE — Détail / éditeur d'une facture
#  NOUVELLE — URL : /factures/<pk>/
# ──────────────────────────────────────────

@login_required
def facture_detail(request, pk):
    """
    Page d'édition d'une facture brouillon.
    Accessible depuis l'onglet Factures du devis.

    """
    facture = get_object_or_404(Facture, pk=pk)
    devis = facture.devis

    # Vérification accès — lecture (comptable inclus pour la validation)
    if not peut_voir_facture(request.user, facture):
        messages.error(request, "Vous n'avez pas accès à cette facture.")
        return redirect('core:devis-list')

    # Factures précédentes validées sur le même devis (pour le récapitulatif)
    # PROTO : on affiche toutes les factures validées, pas les avoirs.
    STATUTS_VALIDES = ('validated', 'sent', 'paid')
    factures_prec = devis.factures.filter(
        status__in=STATUTS_VALIDES,
        type_doc='facture',
    ).exclude(pk=facture.pk).order_by('created_at')

    # Totaux pour le footer
    total_devis = float(devis.total_brut())
    total_deja = float(devis.total_facture())  # méthode existante sur Devis

    return render(request, 'core/facture_detail.html', {
        'facture': facture,
        'devis': devis,
        'factures_prec': factures_prec,
        'total_devis': total_devis,
        'total_deja': total_deja,
        'modifiable': facture.status == 'draft',
    })


# ──────────────────────────────────────────
#  VUE — Aperçu imprimable d'une facture
#  NOUVELLE — URL : /factures/<pk>/apercu/
# ──────────────────────────────────────────

@login_required
def facture_apercu(request, pk):
    """
    Aperçu HTML imprimable de la facture.
    Lecture seule — même données que facture_detail.
    window.print() déclenché par un bouton, CSS @media print intégré.

    # PROTO : remplacé par WeasyPrint en Phase 3.
    # Affiche uniquement les lignes avec quantite > 0.
    """
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_voir_facture(request.user, facture):
        messages.error(request, "Vous n'avez pas accès à cette facture.")
        return redirect('core:devis-list')
    devis = facture.devis

    from .models import ParametresAssociation
    params = ParametresAssociation.get()

    # Historique de facturation par source (pour affichage dans l'aperçu par TITRE)
    if devis is not None:
        deja_par_source = calc_deja_par_source_detail(devis, facture)
        _STATUTS_VALIDES = ('validated', 'sent', 'paid')
        _factures_prec_qs = devis.factures.filter(
            status__in=_STATUTS_VALIDES, type_doc='facture',
        ).exclude(pk=facture.pk).order_by('created_at')
        ref_to_info = {
            f.get_reference(): {
                'date':  f.date_creation.strftime('%d/%m/%Y'),
                'notes': f.notes or '',
            }
            for f in _factures_prec_qs
        }
    else:
        deja_par_source = {}
        ref_to_info     = {}

    def filtrer_lignes(lignes_qs, parent_non_facture=False):
        result = []
        for lf in lignes_qs:
            lf.non_facture = parent_non_facture or float(lf.quantite) == 0
            if lf.type_ligne == 'TITRE':
                if not lf.non_facture:
                    lf.pu_section = lf.total() / lf.quantite
                entry = deja_par_source.get(lf.ligne_devis_source_id) if lf.ligne_devis_source_id else None
                lf.refs_prec = [
                    {
                        'ref':     ref,
                        'montant': montant,
                        'date':    ref_to_info.get(ref, {}).get('date', ''),
                        'notes':   ref_to_info.get(ref, {}).get('notes', ''),
                    }
                    for ref, montant in sorted((entry or {}).get('refs', {}).items())
                    if montant > 0
                ]
                lf.enfants_filtres = filtrer_lignes(lf.enfants.all(), lf.non_facture)
                result.append(lf)
            else:
                result.append(lf)
        return result

    # Pour un avoir, les quantités sont négatives → on garde tout (pas de filtre > 0),
    # mais on attache enfants_filtres pour que les titres affichent leurs enfants.
    def garder_tout(lignes_qs):
        result = []
        for lf in lignes_qs:
            if lf.type_ligne == 'TITRE':
                lf.refs_prec = []
                lf.enfants_filtres = garder_tout(lf.enfants.all())
                result.append(lf)
            else:
                result.append(lf)
        return result

    if facture.type_doc == 'avoir':
        lignes_filtrees = garder_tout(facture.lignes.filter(parent=None).exclude(type_ligne='FIN'))
    else:
        lignes_filtrees = filtrer_lignes(facture.lignes.filter(parent=None).exclude(type_ligne='FIN'))

    from decimal import Decimal
    if devis is not None:
        # Lignes financement du devis (FIN, niveau racine) — section informative
        lignes_fin = list(devis.lignes.filter(parent=None, type_ligne='FIN'))

        # Acomptes déduits uniquement sur la première facture non-acompte du devis
        premiere_facture = devis.factures.exclude(
            type_doc='acompte'
        ).exclude(status='cancelled').order_by('created_at').first()

        if premiere_facture and premiere_facture.pk == facture.pk:
            acomptes = devis.factures.filter(
                type_doc='acompte',
            ).exclude(status='draft').order_by('created_at')
            total_acomptes = sum(
                a.montant for a in acomptes if a.status == 'paid'
            )
        else:
            acomptes = devis.factures.none()
            total_acomptes = Decimal('0')
        coordonnees_cb = facture.coordonnees_cb or devis.coordonnees_cb
    else:
        # Facture compta (sans devis) : pas de financement ni d'acomptes
        lignes_fin = []
        acomptes = Facture.objects.none()
        total_acomptes = Decimal('0')
        coordonnees_cb = facture.coordonnees_cb

    solde = facture.montant - total_acomptes

    ctx = {
        'facture': facture,
        'devis': devis,
        'client': facture.get_client(),
        'contact': facture.contact_client,
        'params': params,
        'lignes': lignes_filtrees,
        'lignes_fin': lignes_fin,
        'coordonnees_cb': coordonnees_cb,
        'acomptes': acomptes,
        'total_acomptes': total_acomptes,
        'solde': solde,
        'has_acomptes': total_acomptes > 0,
    }
    if request.GET.get('download') == '1':
        return _render_pdf(request, 'core/facture_apercu.html', ctx, f'{facture.get_reference()}.pdf')
    return render(request, 'core/facture_apercu.html', ctx)


# ──────────────────────────────────────────
#  VUE — Sauvegarde du libellé inline
#  NOUVELLE — URL : /factures/<pk>/libelle/
# ──────────────────────────────────────────

@login_required
@require_POST
def facture_libelle_save(request, pk):
    """
    Sauvegarde le libellé court d'une facture (éditable inline dans le
    récapitulatif des factures précédentes).

    # PROTO : pas de validation de longueur côté serveur pour l'instant.
    # max_length=200 est géré par le modèle.
    """
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_modifier_facture(request.user, facture):
        return json_error_permission()
    data, err = parse_json_request(request)
    if err:
        return err
    libelle = data.get('libelle', '').strip()[:200]
    facture.libelle = libelle
    facture.save(update_fields=['libelle'])
    return JsonResponse({'ok': True, 'libelle': libelle})

# ══════════════════════════════════════════
#  LIGNES FACTURE (API JSON)
# ══════════════════════════════════════════

@login_required
def lignes_facture_get(request, pk):
    """
    Retourne les lignes de la facture + montant "déjà facturé" par ligne.

    # PROTO : deja_par_source est calculé à chaque appel.
    # Optimisation possible si les factures sont nombreuses.
    """
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_voir_facture(request.user, facture):
        return json_error_permission()
    deja_par_source = calc_deja_par_source_detail(facture.devis, facture)
    racines = facture.lignes.filter(parent=None)
    data = [ligne_facture_to_dict(l, deja_par_source) for l in racines]

    # Factures précédentes pour le récapitulatif JS (libellé + montant global)
    STATUTS_VALIDES = ('validated', 'sent', 'paid')
    factures_prec = facture.devis.factures.filter(
        status__in=STATUTS_VALIDES,
        type_doc='facture',
    ).exclude(pk=facture.pk).order_by('created_at')

    factures_prec_data = [
        {
            'pk': f.pk,
            'reference': f.get_reference(),
            'libelle': f.libelle or f.notes or '',
            'montant': float(f.montant),
            'date': f.date_creation.strftime('%d/%m/%Y'),
            'libelle_save_url': f'/factures/{f.pk}/libelle/',
        }
        for f in factures_prec
    ]

     # Acomptes versés — affichés dans la zone financement
    acomptes = facture.devis.factures.filter(
        status__in=STATUTS_VALIDES,
        type_doc='acompte',
    ).order_by('created_at')

    acomptes_data = [
        {
            'pk': f.pk,
            'reference': f.get_reference(),
            'montant': float(f.montant),
            'notes': f.notes or '',
            'date_versement': f.date_versement.strftime('%d/%m/%Y') if f.date_versement else '',
        }
        for f in acomptes
    ]

    return JsonResponse({
        'lignes': data,
        'montant': float(facture.montant),
        'taux_mo': float(facture.devis.taux_mo),
        'total_devis': float(facture.devis.total_brut()),
        'total_deja': float(facture.devis.total_facture()),
        'factures_prec': factures_prec_data,
        'acomptes': acomptes_data,
    })

@login_required
@require_POST
def lignes_facture_save(request, pk):
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_modifier_facture(request.user, facture):
        return json_error_permission()
    if facture.status != 'draft':
        return json_error('Facture non modifiable')
    data, err = parse_json_request(request)
    if err:
        return err
    notes = data.get('notes', None)
    if notes is not None:
        facture.notes = notes
    lignes = data.get('lignes', [])

    facture.lignes.all().delete()

    create_lignes = build_lignes_creator(
        LigneFacture, {'facture': facture},
        with_quantite_originale=True, with_source=True)
    create_lignes(lignes)

    # Recalculer le montant total de la facture
    total = sum(l.total() for l in facture.lignes.filter(parent=None))
    facture.montant = total
    facture.save()

    return JsonResponse({'ok': True, 'montant': float(total)})


@login_required
def utilisateurs_list(request):
    """
    Liste des utilisateurs.
    - Admin : voit tous les utilisateurs actifs et inactifs
    - Responsable : voit uniquement les membres de ses équipes
    """
    if not peut_gerer_utilisateurs(request.user):
        messages.error(request, 'Accès réservé aux administrateurs et responsables.')
        return redirect('core:dashboard')

    profil = get_profil(request.user)

    if profil.role == 'admin':
        profils = ProfilUtilisateur.objects.select_related(
            'user', 'service__territoire', 'responsable__user'
        ).prefetch_related('equipes__service').order_by(
            'user__last_name', 'user__first_name'
        )
    else:
        # Responsable — uniquement les membres de ses équipes
        equipes_ids = profil.equipes.values_list('pk', flat=True)
        profils = ProfilUtilisateur.objects.filter(
            equipes__pk__in=equipes_ids
        ).exclude(
            pk=profil.pk  # s'exclut lui-même de la liste
        ).select_related(
            'user', 'service__territoire', 'responsable__user'
        ).prefetch_related('equipes__service').distinct().order_by(
            'user__last_name', 'user__first_name'
        )

    return render(request, 'core/utilisateurs_list.html', {
        'profils': profils,
    })


@login_required
def utilisateur_create(request):
    """
    Création d'un nouvel utilisateur.
    - Génère un mot de passe temporaire affiché une seule fois
    - Copie optionnelle de la bibliothèque d'un utilisateur existant
    - Admin : peut attribuer tous les rôles et toutes les équipes
    - Responsable : ne peut pas créer d'admin, voit uniquement ses équipes
    """
    if not peut_gerer_utilisateurs(request.user):
        messages.error(request, 'Accès réservé aux administrateurs et responsables.')
        return redirect('core:dashboard')

    profil_courant = get_profil(request.user)

    if request.method == 'POST':
        username   = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        role       = request.POST.get('role', 'technicien')
        service_id = request.POST.get('service') or None
        equipes_ids = request.POST.getlist('equipes')
        responsable_id = request.POST.get('responsable') or None
        copier_biblio_de = request.POST.get('copier_biblio_de') or None

        # Validation
        if not username:
            messages.error(request, 'Le nom d\'utilisateur est obligatoire.')
            return redirect('core:utilisateur-create')

        # TODO (avant passage Phase 4) : limiter les emails à @compagnonsbatisseurs.eu
        # if email and not email.endswith(f'@{DOMAINE_AUTORISE}'):
        #     messages.error(request, f'L\'adresse email doit être une adresse @{DOMAINE_AUTORISE}.')
        #     return redirect('core:utilisateur-create')

        if User.objects.filter(username=username).exists():
            messages.error(request, f'Le nom d\'utilisateur "{username}" est déjà utilisé.')
            return redirect('core:utilisateur-create')

        # Responsable ne peut pas créer d'admin
        if profil_courant.role == 'responsable' and role == 'admin':
            messages.error(request, 'Vous ne pouvez pas créer un administrateur.')
            return redirect('core:utilisateur-create')

        # Génération mot de passe temporaire
        mdp_temp = ''.join(random.choices(
            string.ascii_letters + string.digits, k=12
        ))

        # Création User Django
        user = User.objects.create_user(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            password=mdp_temp,
        )

        # Création ProfilUtilisateur
        nouveau_profil = ProfilUtilisateur.objects.create(
            user=user,
            role=role,
            service_id=service_id,
            responsable_id=responsable_id,
        )

        # Attribution des équipes
        if equipes_ids:
            nouveau_profil.equipes.set(equipes_ids)

        # Copie de bibliothèque
        if copier_biblio_de:
            try:
                source_user = User.objects.get(pk=copier_biblio_de)
                biblio_source, _ = Bibliotheque.objects.get_or_create(user=source_user)
                Bibliotheque.objects.create(
                    user=user,
                    lignes=biblio_source.lignes,
                )
            except User.DoesNotExist:
                pass
        else:
            Bibliotheque.objects.create(user=user, lignes=[])

        # Envoi de l'email d'invitation (best-effort) + affichage du mot de passe.
        # BETA : la délivrance vers les adresses @compagnonsbatisseurs.eu n'est pas
        # fiable (M365 rejette les mails relayés par Brevo, faute d'authentification
        # DNS du domaine — voir NOTES_DEV § Infra). On affiche donc TOUJOURS le mot
        # de passe temporaire pour communication manuelle, et l'email reste un bonus
        # (utile pour d'éventuelles adresses externes).
        nom_affiche = f'{first_name} {last_name}'.strip() or username
        email_statut = None

        # Texte d'invitation construit une seule fois : sert à l'envoi ET à
        # l'affichage à l'écran (copier-coller pour communication manuelle).
        message_invitation = (
            f'Bonjour {first_name or username},\n\n'
            f'Votre compte CB Bretagne a été créé.\n\n'
            f'Identifiant : {username}\n'
            f'Mot de passe temporaire : {mdp_temp}\n\n'
            f'Connectez-vous ici : {settings.SITE_URL}/login/\n\n'
            f'Merci de changer votre mot de passe dès votre première connexion.\n\n'
            f'Un manuel utilisateur est disponible directement sur le site : {settings.SITE_URL}/aide/\n\n'
            f'CB Bretagne'
        )

        if email:
            try:
                send_mail(
                    subject='Votre compte CB Bretagne',
                    message=message_invitation,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
                nouveau_profil.invitation_envoyee = True
                nouveau_profil.save(update_fields=['invitation_envoyee'])
                email_statut = (
                    f"Un email d'invitation a été envoyé à {email}. ⚠️ En beta, la "
                    f"délivrance vers les adresses @compagnonsbatisseurs.eu n'est pas "
                    f"garantie : copiez le message ci-dessous et envoyez-le par un autre "
                    f"canal (Teams, votre messagerie)."
                )
            except Exception as e:
                logger.error('Invitation email error (user %s): %s', username, e)
                email_statut = (
                    f"L'email d'invitation vers {email} n'a pas pu être envoyé — "
                    f"copiez le message ci-dessous et transmettez-le manuellement."
                )
        else:
            email_statut = 'Aucune adresse email renseignée — copiez le message ci-dessous et transmettez-le manuellement.'

        # Toujours afficher le message d'invitation complet (une seule fois)
        request.session['nouveau_user_mdp'] = {
            'username': username,
            'mdp': mdp_temp,
            'nom': nom_affiche,
            'email': email,
            'message': message_invitation,
            'email_statut': email_statut,
        }
        return redirect('core:utilisateur-create-succes')

    # GET — construction du formulaire
    if profil_courant.role == 'admin':
        equipes_disponibles = Equipe.objects.select_related('service__territoire').all().order_by('service__nom', 'nom')
        roles_disponibles = ProfilUtilisateur.ROLE_CHOICES
    else:
        equipes_disponibles = profil_courant.equipes.select_related('service__territoire').all().order_by('service__nom', 'nom')
        # Responsable ne peut pas créer d'admin
        roles_disponibles = [
            (k, v) for k, v in ProfilUtilisateur.ROLE_CHOICES if k != 'admin'
        ]

    return render(request, 'core/utilisateur_form.html', {
        'equipes': equipes_disponibles,
        'roles': roles_disponibles,
        'services': Service.objects.select_related('territoire').all(),
        'responsables': ProfilUtilisateur.objects.filter(
            role__in=('admin', 'responsable'), user__is_active=True
        ).select_related('user').order_by('user__last_name'),
        'utilisateurs_biblio': User.objects.filter(
            is_active=True
        ).select_related('profil').order_by('last_name', 'first_name'),
        'mode': 'creation',
    })


@login_required
def utilisateur_create_succes(request):
    """
    Page affichée une seule fois après la création d'un utilisateur.
    Affiche le mot de passe temporaire puis le supprime de la session.
    """
    if not peut_gerer_utilisateurs(request.user):
        return redirect('core:dashboard')

    infos = request.session.pop('nouveau_user_mdp', None)
    if not infos:
        return redirect('core:utilisateurs-list')

    return render(request, 'core/utilisateur_succes.html', {
        'infos': infos,
    })


@login_required
def utilisateur_edit(request, pk):
    """
    Modification d'un utilisateur existant : rôle, équipes, service, responsable.
    Ne modifie pas le mot de passe (feature séparée).
    """
    if not peut_gerer_utilisateurs(request.user):
        messages.error(request, 'Accès réservé aux administrateurs et responsables.')
        return redirect('core:dashboard')

    cible_profil = get_object_or_404(ProfilUtilisateur, user__pk=pk)
    profil_courant = get_profil(request.user)

    if not peut_gerer_cet_utilisateur(request.user, cible_profil):
        messages.error(request, 'Vous ne pouvez pas modifier cet utilisateur.')
        return redirect('core:utilisateurs-list')

    if request.method == 'POST':
        # Infos de base
        cible_profil.user.first_name = request.POST.get('first_name', '').strip()
        cible_profil.user.last_name  = request.POST.get('last_name', '').strip()
        cible_profil.user.email      = request.POST.get('email', '').strip()
        cible_profil.user.save()

        role = request.POST.get('role', cible_profil.role)

        # Responsable ne peut pas promouvoir en admin
        if profil_courant.role == 'responsable' and role == 'admin':
            messages.error(request, 'Vous ne pouvez pas attribuer le rôle administrateur.')
            return redirect('core:utilisateur-edit', pk=pk)

        cible_profil.role          = role
        cible_profil.service_id    = request.POST.get('service') or None
        cible_profil.responsable_id = request.POST.get('responsable') or None
        cible_profil.save()

        equipes_ids = request.POST.getlist('equipes')
        cible_profil.equipes.set(equipes_ids)

        messages.success(request, f'Utilisateur "{cible_profil.user.username}" mis à jour.')
        return redirect('core:utilisateurs-list')

    # GET
    if profil_courant.role == 'admin':
        equipes_disponibles = Equipe.objects.select_related('service__territoire').all().order_by('service__nom', 'nom')
        roles_disponibles = ProfilUtilisateur.ROLE_CHOICES
    else:
        equipes_disponibles = profil_courant.equipes.select_related('service__territoire').all().order_by('service__nom', 'nom')
        roles_disponibles = [
            (k, v) for k, v in ProfilUtilisateur.ROLE_CHOICES if k != 'admin'
        ]

    return render(request, 'core/utilisateur_form.html', {
        'cible': cible_profil,
        'equipes': equipes_disponibles,
        'roles': roles_disponibles,
        'services': Service.objects.select_related('territoire').all(),
        'responsables': ProfilUtilisateur.objects.filter(
            role__in=('admin', 'responsable'), user__is_active=True
        ).select_related('user').order_by('user__last_name'),
        'mode': 'edition',
    })


@login_required
@require_POST
def utilisateur_toggle(request, pk):
    """
    Bascule is_active de l'utilisateur (désactiver / réactiver).
    Un admin ne peut pas se désactiver lui-même.
    """
    if not peut_gerer_utilisateurs(request.user):
        messages.error(request, 'Accès réservé aux administrateurs et responsables.')
        return redirect('core:dashboard')

    cible_profil = get_object_or_404(ProfilUtilisateur, user__pk=pk)

    if not peut_gerer_cet_utilisateur(request.user, cible_profil):
        messages.error(request, 'Vous ne pouvez pas modifier cet utilisateur.')
        return redirect('core:utilisateurs-list')

    # Empêcher l'auto-désactivation
    if cible_profil.user == request.user:
        messages.error(request, 'Vous ne pouvez pas désactiver votre propre compte.')
        return redirect('core:utilisateurs-list')

    cible_user = cible_profil.user
    cible_user.is_active = not cible_user.is_active
    cible_user.save()

    action = 'réactivé' if cible_user.is_active else 'désactivé'
    messages.success(request, f'Utilisateur "{cible_user.username}" {action}.')
    return redirect('core:utilisateurs-list')


# ══════════════════════════════════════════
#  OUTILS COMPTA — Factures structure / Appels de convention / Avoirs
# ══════════════════════════════════════════
#
# Factures créées directement, sans devis. Réservées aux rôles compta
# (peut_acceder_compta). Réutilisent le modèle Facture/LigneFacture :
#   - devis = None
#   - type_doc in ('structure', 'appel')  (+ 'avoir' généré depuis une facture)
#   - lignes plates à 2 niveaux : TITRE (groupe) + F (forfait)
# La validation suit peut_valider_facture (admin ou comptable). draft = proforma.

from django.db.models import Q

# Métadonnées d'affichage par type d'outil compta
COMPTA_TYPES = {
    'structure': {'titre': 'Factures Structure',     'singulier': 'Facture structure',
                  'list_url': 'core:compta-structures-list'},
    'appel':     {'titre': 'Appels de convention',   'singulier': 'Appel de convention',
                  'list_url': 'core:compta-appels-list'},
}


def _compta_lignes_to_dict(ligne):
    """Sérialise une LigneFacture compta (TITRE/forfait) en dict JSON."""
    return {
        'id': ligne.pk,
        'type_ligne': ligne.type_ligne,
        'description': ligne.description,
        'quantite': float(ligne.quantite),
        'unite': ligne.unite,
        'cout_unitaire': float(ligne.cout_unitaire) if ligne.cout_unitaire is not None else None,
        'ordre': ligne.ordre,
        'enfants': [_compta_lignes_to_dict(e) for e in ligne.enfants.all().order_by('ordre')],
    }


@login_required
def factures_compta_list(request, type_doc):
    """Liste des factures d'un outil compta + leurs avoirs liés."""
    if not peut_acceder_compta(request.user):
        messages.error(request, "Accès réservé à la comptabilité.")
        return redirect('core:dashboard')

    meta = COMPTA_TYPES[type_doc]
    factures = Facture.objects.filter(
        Q(type_doc=type_doc, devis__isnull=True)
        | Q(type_doc='avoir', facture_origine__type_doc=type_doc)
    ).select_related('client', 'facture_origine', 'created_by__profil').prefetch_related('avoirs').order_by('-created_at')

    q = request.GET.get('q', '').strip()
    auteur_id = request.GET.get('auteur', '')
    if q:
        factures = factures.filter(
            Q(numero__icontains=q) | Q(notes__icontains=q)
            | Q(client__nom__icontains=q) | Q(destinataire__icontains=q)
        )
    if auteur_id:
        factures = factures.filter(created_by_id=auteur_id)

    page_obj, base_qs = paginer(request, factures)

    return render(request, 'core/facture_compta_list.html', {
        'factures': page_obj,
        'page_obj': page_obj,
        'base_qs': base_qs,
        'type_doc': type_doc,
        'meta': meta,
        'auteurs': User.objects.filter(
            factures_creees__type_doc__in=[type_doc, 'avoir']
        ).distinct().order_by('first_name', 'username'),
        'auteur_filter': auteur_id,
        'q': q,
    })


@login_required
def facture_compta_create(request, type_doc):
    """Création directe d'une facture compta (structure ou appel)."""
    if not peut_acceder_compta(request.user):
        messages.error(request, "Accès réservé à la comptabilité.")
        return redirect('core:dashboard')

    meta = COMPTA_TYPES[type_doc]
    profil = get_profil(request.user)

    if request.method == 'POST':
        client_id = request.POST.get('client') or None
        client = Client.objects.filter(pk=client_id).first() if client_id else None
        contact_id = request.POST.get('contact_client') or None
        contact = None
        if contact_id and client:
            contact = ContactClient.objects.filter(pk=contact_id, client=client).first()

        objet = request.POST.get('notes', '').strip()
        try:
            echeance_jours = int(request.POST.get('echeance_jours') or 30)
        except (TypeError, ValueError):
            echeance_jours = 30

        destinataire = (client.nom if client else request.POST.get('destinataire', '').strip())

        facture = Facture.objects.create(
            type_doc=type_doc,
            devis=None,
            client=client,
            contact_client=contact,
            destinataire=destinataire or 'Sans nom',
            notes=objet,
            date_echeance=timezone.localdate() + timedelta(days=echeance_jours),
            conditions_facture=profil.conditions_facture,
            coordonnees_cb=profil.coordonnees_cb,
            created_by=request.user,
        )
        add_audit(request.user, f"Création {meta['singulier']} {facture.get_reference()}",
                  facture=facture)
        return redirect('core:compta-facture-detail', pk=facture.pk)

    return render(request, 'core/facture_compta_form.html', {
        'type_doc': type_doc,
        'meta': meta,
    })


@login_required
def facture_compta_detail(request, pk):
    """Éditeur 2 niveaux (titres + forfaits) d'une facture compta ou d'un avoir."""
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_voir_facture(request.user, facture):
        messages.error(request, "Vous n'avez pas accès à cette facture.")
        return redirect('core:dashboard')

    meta = COMPTA_TYPES.get(
        facture.facture_origine.type_doc if facture.type_doc == 'avoir' and facture.facture_origine
        else facture.type_doc,
        {'titre': 'Facture', 'singulier': 'Facture', 'list_url': 'core:factures-list'},
    )
    return render(request, 'core/facture_compta_detail.html', {
        'facture': facture,
        'meta': meta,
        'modifiable': facture.status == 'draft' and peut_modifier_facture(request.user, facture),
    })


@login_required
def lignes_compta_get(request, pk):
    """Retourne les lignes (titres + forfaits) + montant de la facture compta."""
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_voir_facture(request.user, facture):
        return json_error_permission()
    racines = facture.lignes.filter(parent=None).order_by('ordre')
    return JsonResponse({
        'lignes': [_compta_lignes_to_dict(l) for l in racines],
        'montant': float(facture.montant),
        'notes': facture.notes,
    })


@login_required
@require_POST
def lignes_compta_save(request, pk):
    """Remplace les lignes (titres + forfaits) ; recalcule le montant (± pour les avoirs)."""
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_modifier_facture(request.user, facture):
        return json_error_permission()
    if facture.status != 'draft':
        return json_error('Facture non modifiable')
    data, err = parse_json_request(request)
    if err:
        return err

    notes = data.get('notes', None)
    if notes is not None:
        facture.notes = notes
    lignes = data.get('lignes', [])

    facture.lignes.all().delete()

    create_lignes = build_lignes_creator(
        LigneFacture, {'facture': facture},
        with_ouvert=False, recurse_only_titre=True)
    create_lignes(lignes)

    total = sum(l.total() for l in facture.lignes.filter(parent=None))
    facture.montant = total
    facture.save()
    return JsonResponse({'ok': True, 'montant': float(total)})


@login_required
@require_POST
def facture_compta_valider(request, pk):
    """Validation comptable d'une facture compta — assigne le numéro."""
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_acceder_compta(request.user) or not peut_valider_facture(request.user, facture):
        messages.error(request, "Validation réservée à la comptabilité.")
        return redirect('core:compta-facture-detail', pk=pk)
    if not facture.numero:
        facture.numero = gen_numero_facture(facture.type_doc)
    facture.status = 'validated'
    facture.validated_by = request.user
    facture.validated_at = timezone.now()
    facture.save()
    add_audit(request.user, f"Validation {facture.numero} ({facture.destinataire})", facture=facture)
    return redirect('core:compta-facture-detail', pk=pk)


@login_required
@require_POST
def facture_compta_status(request, pk):
    """Changement de statut d'une facture compta (validée → envoyée → payée)."""
    facture = get_object_or_404(Facture, pk=pk)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'
    if not peut_modifier_facture(request.user, facture):
        if is_ajax:
            return json_error_permission()
        messages.error(request, 'Vous ne pouvez pas modifier cette facture.')
        return redirect('core:compta-facture-detail', pk=pk)
    old = facture.status
    facture.status = request.POST.get('status', facture.status)
    facture.save()
    add_audit(request.user, f"Statut {facture.get_reference()} : {old} → {facture.status}", facture=facture)
    if is_ajax:
        return JsonResponse({'ok': True})
    return redirect('core:compta-facture-detail', pk=pk)


@login_required
@require_POST
def facture_compta_delete(request, pk):
    """Suppression d'une facture compta — brouillons uniquement."""
    facture = get_object_or_404(Facture, pk=pk)
    if not peut_supprimer_facture(request.user, facture):
        messages.error(request, 'Vous ne pouvez pas supprimer cette facture.')
        return redirect('core:compta-facture-detail', pk=pk)
    # Liste de retour selon le type (avoir → liste de la facture d'origine)
    type_retour = (facture.facture_origine.type_doc
                   if facture.type_doc == 'avoir' and facture.facture_origine
                   else facture.type_doc)
    list_url = COMPTA_TYPES.get(type_retour, {}).get('list_url', 'core:dashboard')
    ref = facture.get_reference()
    add_audit(request.user, f"Suppression facture compta {ref}", facture=None)
    facture.delete()
    messages.success(request, f'Facture {ref} supprimée.')
    return redirect(list_url)


@login_required
@require_POST
def facture_compta_duplicate(request, pk):
    """Duplique une facture compta (structure/appel) en un nouveau brouillon."""
    source = get_object_or_404(Facture, pk=pk)
    if not peut_acceder_compta(request.user):
        messages.error(request, "Accès réservé à la comptabilité.")
        return redirect('core:dashboard')
    if source.type_doc not in ('structure', 'appel'):
        messages.error(request, "Seules les factures structure et appels peuvent être dupliqués.")
        return redirect('core:compta-facture-detail', pk=source.pk)

    copie = Facture.objects.create(
        type_doc=source.type_doc,
        devis=None,
        client=source.client,
        contact_client=source.contact_client,
        destinataire=source.destinataire,
        notes=source.notes,
        date_echeance=source.date_echeance,
        conditions_facture=source.conditions_facture,
        coordonnees_cb=source.coordonnees_cb,
        created_by=request.user,
    )

    def copier(lignes_src, parent=None):
        for l in lignes_src.order_by('ordre'):
            nl = LigneFacture.objects.create(
                facture=copie, parent=parent,
                type_ligne=l.type_ligne,
                description=l.description,
                quantite=l.quantite,
                quantite_originale=l.quantite_originale,
                unite=l.unite,
                cout_unitaire=l.cout_unitaire,
                ordre=l.ordre,
            )
            copier(l.enfants.all(), parent=nl)

    copier(source.lignes.filter(parent=None))
    copie.montant = sum(l.total() for l in copie.lignes.filter(parent=None))
    copie.save()

    add_audit(request.user, f"Duplication {source.get_reference()} → brouillon", facture=copie)
    return redirect('core:compta-facture-detail', pk=copie.pk)


@login_required
@require_POST
def avoir_create(request, facture_pk):
    """
    Crée un avoir depuis une facture validée : copie les lignes avec quantités
    inversées (négatives), modifiables ensuite dans l'éditeur. Fonctionne pour
    tous les types (chantier / structure / appel).
    """
    source = get_object_or_404(Facture, pk=facture_pk)
    if not peut_modifier_facture(request.user, source):
        messages.error(request, "Vous ne pouvez pas créer d'avoir sur cette facture.")
        return redirect('core:dashboard')
    if source.status not in ('validated', 'sent', 'paid'):
        messages.error(request, "Un avoir ne peut être créé que depuis une facture validée.")
        return redirect('core:compta-facture-detail', pk=source.pk)

    avoir = Facture.objects.create(
        type_doc='avoir',
        devis=source.devis,
        client=source.client,
        contact_client=source.contact_client,
        facture_origine=source,
        destinataire=source.destinataire,
        notes=f"Avoir sur facture {source.get_reference()}",
        conditions_facture=source.conditions_facture,
        coordonnees_cb=source.coordonnees_cb,
        created_by=request.user,
    )

    # Copie des lignes avec quantités inversées (récursif : titres + enfants)
    def copier_negatif(lignes_src, parent=None):
        for l in lignes_src.order_by('ordre'):
            qte = l.quantite if l.type_ligne == 'TITRE' else (l.quantite * Decimal('-1'))
            nl = LigneFacture.objects.create(
                facture=avoir, parent=parent,
                type_ligne=l.type_ligne,
                description=l.description,
                quantite=qte,
                quantite_originale=l.quantite_originale,
                unite=l.unite,
                cout_unitaire=l.cout_unitaire,
                ordre=l.ordre,
                ligne_devis_source=l.ligne_devis_source,
            )
            copier_negatif(l.enfants.all(), parent=nl)

    copier_negatif(source.lignes.filter(parent=None))
    avoir.montant = sum(l.total() for l in avoir.lignes.filter(parent=None))
    avoir.save()

    add_audit(request.user, f"Création avoir sur {source.get_reference()}",
              devis=source.devis, facture=avoir)
    return redirect('core:compta-facture-detail', pk=avoir.pk)


# ── Contacts client (carnet optionnel) ──────────────────────────────

@login_required
def client_contacts_get(request, client_pk):
    """Liste JSON des contacts d'un client (pour le select du formulaire compta)."""
    if not peut_acceder_compta(request.user):
        return json_error_permission()
    client = get_object_or_404(Client, pk=client_pk)
    contacts = [
        {'id': c.pk, 'label': str(c), 'service': c.service, 'nom': c.nom,
         'email': c.email, 'telephone': c.telephone}
        for c in client.contacts.all()
    ]
    return JsonResponse({'contacts': contacts})


@login_required
@require_POST
def contact_client_create(request):
    """Ajout rapide d'un contact à un client → {id, label}."""
    data, err = parse_json_request(request)
    if err:
        return err
    client = get_object_or_404(Client, pk=data.get('client'))
    service = (data.get('service') or '').strip()
    nom = (data.get('nom') or '').strip()
    if not service and not nom:
        return json_error('Service ou nom requis')
    contact = ContactClient.objects.create(
        client=client,
        service=service,
        nom=nom,
        fonction=(data.get('fonction') or '').strip(),
        email=(data.get('email') or '').strip(),
        telephone=(data.get('telephone') or '').strip(),
    )
    return JsonResponse({'ok': True, 'id': contact.pk, 'label': str(contact)})


@login_required
@require_POST
def contact_client_delete(request, pk):
    """Suppression d'un contact (admin uniquement, comme l'édition client)."""
    if not is_admin(request.user):
        return json_error("Réservé à l'administrateur", status=403)
    contact = get_object_or_404(ContactClient, pk=pk)
    contact.delete()
    return JsonResponse({'ok': True})


